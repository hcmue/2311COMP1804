from openpyxl import Workbook

wb = Workbook() # đã có 1 sheet default: Sheet 1
# Tạo worksheet có tên NIIE
ws = wb.create_sheet("CNTT", 1)
ws["A1"] = "Danh sách chưa tốt nghiệp"	# Ghi ô "A1"
ws.append([2, 3, 4])			# Thêm dòng mới

row_1 = ["Mã số sv",	"Họ lót",	"Tên",	"Giới tính", 
            "Ngày sinh",	"Nơi sinh",	"Chuẩn xét",
            "Điểm TB Tích lũy",	"Số TCBB",	"Số TCTC",	"Số TC học lại"]
for idx in range(len(row_1)):
    ws.cell(row=2, column=idx+1).value = row_1[idx]

wb.save("DSChuaTotNghiep.xlsx")		# Lưu file
