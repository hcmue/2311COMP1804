from tkinter import *
from tkinter.messagebox import showinfo
from openpyxl import Workbook

# B1: Khai báo
root = Tk()
root.geometry("400x300")
root.title("MAIN PAGE")


# B2: Gắn widget (Entry, Text, Label, Button, Combobox, Radiobutton, Checkbox,
# Treeview, Menu)
# các kiểu layout (gắn/đặt widget vô UI):
#   .pack() # Mỗi widget trên 1 dòng
#   .place(x=v, y=v)
#   .grid(row=x,column=x)
# B3: Gắn sự kiện
Label(root, text="DEMO").pack()
Button(root, text="Click me", command=lambda: showinfo("D", "HELLO")).pack()

def export_excel():    
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")

Button(root, text="Export Excel", command=export_excel).pack()

root.mainloop()